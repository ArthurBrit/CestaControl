from datetime import date
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _period_label(start_date: date | None, end_date: date | None) -> str:
    if start_date and end_date:
        return f"Periodo: {start_date:%d/%m/%Y} a {end_date:%d/%m/%Y}"
    if start_date:
        return f"A partir de {start_date:%d/%m/%Y}"
    if end_date:
        return f"Ate {end_date:%d/%m/%Y}"
    return "Todo o historico"


def build_excel_report(rows: Iterable, start_date: date | None, end_date: date | None) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatorio"
    sheet.append(["CestaControl", "", ""])
    sheet.append([_period_label(start_date, end_date), "", ""])
    sheet.append([])
    sheet.append(["Tecnico", "Item", "Total retirado"])

    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14)
    for cell in sheet[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")

    for row in rows:
        sheet.append([row.technician_name, row.item_name, int(row.total or 0)])

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 16

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_pdf_report(rows: Iterable, start_date: date | None, end_date: date | None) -> BytesIO:
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, title="Relatorio CestaControl")
    styles = getSampleStyleSheet()
    story = [
        Paragraph("CestaControl", styles["Title"]),
        Paragraph("Relatorio de retiradas por tecnico", styles["Heading2"]),
        Paragraph(_period_label(start_date, end_date), styles["Normal"]),
        Spacer(1, 16),
    ]

    data = [["Tecnico", "Item", "Total retirado"]]
    data.extend([[row.technician_name, row.item_name, int(row.total or 0)] for row in rows])
    if len(data) == 1:
        data.append(["Sem registros para o filtro selecionado.", "", ""])

    table = Table(data, colWidths=[180, 180, 100])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9ead3")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1f2933")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c8d1c8")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7faf7")]),
            ]
        )
    )
    story.append(table)
    document.build(story)
    output.seek(0)
    return output
